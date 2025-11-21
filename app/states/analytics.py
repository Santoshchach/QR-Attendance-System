import reflex as rx
from sqlmodel import select, func, desc
from datetime import datetime, timedelta
from app.models import Session, Attendance, User
from app.states.auth import AuthState
import random
import string
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill


class AnalyticsState(rx.State):
    """Handle analytics data aggregation and exports."""

    total_sessions: int = 0
    total_students: int = 0
    avg_attendance: float = 0.0
    active_sessions_count: int = 0
    attendance_trends: list[dict] = []
    session_performance: list[dict] = []
    student_distribution: list[dict] = []
    date_range: str = "all"
    selected_course_id: str = "all"
    available_courses: list[dict] = []

    @rx.event
    async def load_stats(self):
        """Load all statistics for the dashboard."""
        auth_state = await self.get_state(AuthState)
        if not auth_state.is_authenticated or auth_state.user_role != "teacher":
            return
        teacher_id = auth_state.user_id
        with rx.session() as session:
            query = select(Session).where(Session.teacher_id == teacher_id)
            if self.selected_course_id != "all":
                query = query.where(Session.id == int(self.selected_course_id))
            if self.date_range == "week":
                week_ago = datetime.now() - timedelta(days=7)
                query = query.where(Session.created_at >= week_ago)
            elif self.date_range == "month":
                month_ago = datetime.now() - timedelta(days=30)
                query = query.where(Session.created_at >= month_ago)
            sessions = session.exec(query).all()
            session_ids = [s.id for s in sessions]
            self.total_sessions = len(sessions)
            self.active_sessions_count = sum((1 for s in sessions if s.is_active))
            if session_ids:
                att_query = select(Attendance).where(
                    Attendance.session_id.in_(session_ids)
                )
                attendances = session.exec(att_query).all()
            else:
                attendances = []
            unique_student_ids = {a.student_id for a in attendances}
            self.total_students = len(unique_student_ids)
            total_attendance_count = len(attendances)
            if self.total_sessions > 0:
                self.avg_attendance = round(
                    total_attendance_count / self.total_sessions, 1
                )
            else:
                self.avg_attendance = 0.0
            date_groups = {}
            for att in attendances:
                date_str = att.scanned_at.strftime("%Y-%m-%d")
                date_groups[date_str] = date_groups.get(date_str, 0) + 1
            self.attendance_trends = [
                {"date": k, "count": v} for k, v in sorted(date_groups.items())
            ]
            sess_groups = {}
            for att in attendances:
                sess_groups[att.session_id] = sess_groups.get(att.session_id, 0) + 1
            self.session_performance = []
            for s in sessions:
                count = sess_groups.get(s.id, 0)
                name = s.course_name
                if len(name) > 15:
                    name = name[:12] + "..."
                self.session_performance.append(
                    {"name": name, "attendees": count, "full_name": s.course_name}
                )
            self.session_performance.reverse()
            student_counts = {}
            for att in attendances:
                student_counts[att.student_id] = (
                    student_counts.get(att.student_id, 0) + 1
                )
            participation = {"High": 0, "Medium": 0, "Low": 0}
            for s_id, count in student_counts.items():
                ratio = count / self.total_sessions if self.total_sessions > 0 else 0
                if ratio >= 0.75:
                    participation["High"] += 1
                elif ratio >= 0.4:
                    participation["Medium"] += 1
                else:
                    participation["Low"] += 1
            self.student_distribution = [
                {
                    "name": "High (>75%)",
                    "value": participation["High"],
                    "fill": "#10b981",
                },
                {
                    "name": "Medium (40-75%)",
                    "value": participation["Medium"],
                    "fill": "#f59e0b",
                },
                {
                    "name": "Low (<40%)",
                    "value": participation["Low"],
                    "fill": "#ef4444",
                },
            ]
            self.student_distribution = [
                x for x in self.student_distribution if x["value"] > 0
            ]
            all_sessions = session.exec(
                select(Session).where(Session.teacher_id == teacher_id)
            ).all()
            courses = {}
            for s in all_sessions:
                courses[s.course_name] = s.id
            self.available_courses = [
                {
                    "label": f"{s.course_name} ({s.created_at.strftime('%m/%d')})",
                    "value": str(s.id),
                }
                for s in all_sessions[-20:]
            ]

    @rx.event
    def set_date_range(self, val: str):
        self.date_range = val
        return AnalyticsState.load_stats

    @rx.event
    def set_course_filter(self, val: str):
        self.selected_course_id = val
        return AnalyticsState.load_stats

    @rx.event
    async def export_pdf(self):
        """Generate and download PDF report."""
        auth_state = await self.get_state(AuthState)
        teacher_id = auth_state.user_id
        with rx.session() as db_session:
            query = select(Session).where(Session.teacher_id == teacher_id)
            if self.selected_course_id != "all":
                query = query.where(Session.id == int(self.selected_course_id))
            sessions = db_session.exec(query).all()
            session_ids = [s.id for s in sessions]
            if not session_ids:
                attendances = []
            else:
                stmt = (
                    select(Attendance, User)
                    .where(Attendance.session_id.in_(session_ids))
                    .where(Attendance.student_id == User.id)
                )
                results = db_session.exec(stmt).all()
                attendances = [(a, u) for a, u in results]
        upload_dir = rx.get_upload_dir()
        upload_dir.mkdir(parents=True, exist_ok=True)
        filename = f"report_{random.randint(1000, 9999)}.pdf"
        filepath = upload_dir / filename
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph("Attendance Report", styles["Title"]))
        elements.append(Spacer(1, 12))
        elements.append(
            Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 24))
        data = [["Session", "Date", "Student Name", "Student ID", "Status"]]
        sess_map = {s.id: (s.course_name, s.created_at) for s in sessions}
        for att, user in attendances:
            s_name, s_date = sess_map.get(att.session_id, ("Unknown", datetime.now()))
            data.append(
                [
                    s_name,
                    s_date.strftime("%Y-%m-%d"),
                    user.full_name,
                    user.student_id,
                    att.status,
                ]
            )
        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)
        return rx.download(url=f"/_upload/{filename}")

    @rx.event
    async def export_excel(self):
        """Generate and download Excel report."""
        auth_state = await self.get_state(AuthState)
        teacher_id = auth_state.user_id
        with rx.session() as db_session:
            query = select(Session).where(Session.teacher_id == teacher_id)
            if self.selected_course_id != "all":
                query = query.where(Session.id == int(self.selected_course_id))
            sessions = db_session.exec(query).all()
            session_ids = [s.id for s in sessions]
            if not session_ids:
                results = []
            else:
                stmt = (
                    select(Attendance, User)
                    .where(Attendance.session_id.in_(session_ids))
                    .where(Attendance.student_id == User.id)
                )
                results = db_session.exec(stmt).all()
        upload_dir = rx.get_upload_dir()
        upload_dir.mkdir(parents=True, exist_ok=True)
        filename = f"export_{random.randint(1000, 9999)}.xlsx"
        filepath = upload_dir / filename
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Data"
        headers = [
            "Session Name",
            "Session Date",
            "Student Name",
            "Student ID",
            "Scanned At",
            "Status",
        ]
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        sess_map = {s.id: (s.course_name, s.created_at) for s in sessions}
        for att, user in results:
            s_name, s_date = sess_map.get(att.session_id, ("Unknown", datetime.now()))
            ws.append(
                [
                    s_name,
                    s_date.strftime("%Y-%m-%d"),
                    user.full_name,
                    user.student_id,
                    att.scanned_at.strftime("%Y-%m-%d %H:%M:%S"),
                    att.status,
                ]
            )
        wb.save(filepath)
        return rx.download(url=f"/_upload/{filename}")